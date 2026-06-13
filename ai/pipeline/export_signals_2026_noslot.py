#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline/export_signals_2026_noslot.py

Export tất cả signal 01/03/2026 → 22/05/2026 không giới hạn slot.

Filters áp dụng (giống live system):
  T1: prob >= base_cutoff (dynamic regime-aware)
  T2: vol_ratio_20d >= 0.70
  T3: không có bearish candles (shooting_star / engulfing / evening_star)
  HC: hardcheck_pass == 1
  Sector cutoffs: per-sector override (BĐS 0.58, etc.)
  Regime: deep_bear → SECTOR_CUTOFFS_DEEP_BEAR (BĐS override OFF)
  Skip recovery regime (giống live system)

Output Excel:
  📊 Tổng kết   — stats tổng hợp + theo tháng + theo sector
  📋 All Signal — tất cả signal đã qua filter (NO slot cap)
  🔍 Regime     — breakdown theo regime

Usage:
    python -m pipeline.export_signals_2026_noslot
"""
from __future__ import annotations

import pickle
import sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT     = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

MODEL_PATH  = ROOT / "data/models/lgbm_model_20d.pkl"
TRAIN_SET   = ROOT / "data/labels/training_set_20d.parquet"
OHLCV_DIR   = ROOT / "data/ohlcv"
OUTPUT_PATH = ROOT / "data/exports/signal_detail_2026_Mar_May_noslot.xlsx"
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

DATE_START  = "2026-03-01"
DATE_END    = "2026-05-22"
BASE_CUTOFF = 0.52
HOLD_DAYS   = 20
CAPITAL_PER = 100_000_000
WIN_THRESH  = 5.0      # % for label=WIN
SL_THRESH   = -10.0   # % stop-loss (label definition)
FEAT_DIR    = ROOT / "data" / "features"


# ── price helpers ──────────────────────────────────────────────────────────────

def load_price_map() -> dict[str, pd.DataFrame]:
    pm = {}
    for f in OHLCV_DIR.glob("*.parquet"):
        try:
            df = pd.read_parquet(f)[["time", "close"]]
            df["time"] = pd.to_datetime(df["time"])
            pm[f.stem] = df.set_index("time").sort_index()
        except Exception:
            pass
    return pm


def get_calendar(pm: dict) -> list[pd.Timestamp]:
    s: set = set()
    for df in pm.values():
        s.update(df.index.tolist())
    return sorted(s)


def get_price(sym: str, date, pm: dict):
    if date is None or sym not in pm:
        return None
    df = pm[sym]
    if date in df.index:
        return float(df.loc[date, "close"])
    for d in [1, -1, 2, -2]:
        alt = date + pd.Timedelta(days=d)
        if alt in df.index:
            return float(df.loc[alt, "close"])
    return None


def exit_date(entry: pd.Timestamp, cal: list, hold: int):
    try:
        idx = cal.index(entry)
        ei = idx + hold
        return cal[ei] if ei < len(cal) else None
    except (ValueError, IndexError):
        return None


# ── model ──────────────────────────────────────────────────────────────────────

def load_model():
    obj = pickle.load(open(MODEL_PATH, "rb"))
    return (obj["model"], obj["feat_cols"]) if isinstance(obj, dict) else (obj, [])


# ── live filters (same as backtest_live_system.apply_live_filters) ─────────────

def apply_live_filters(df: pd.DataFrame, prob: np.ndarray,
                       cutoff: float = BASE_CUTOFF) -> pd.Series:
    idx = df.index
    t1  = pd.Series(prob >= cutoff, index=idx)

    vol = df.get("vol_ratio_20d", pd.Series(np.nan, index=idx))
    t2  = t1 & ~(vol.isna() | (vol >= 0.70))   # fail if vol < 0.70
    t2_pass = t1 & (vol.isna() | (vol >= 0.70))

    ss  = df.get("candle_shooting_star",     pd.Series(0, index=idx)).fillna(0)
    be  = df.get("candle_bearish_engulfing", pd.Series(0, index=idx)).fillna(0)
    es  = df.get("candle_evening_star",      pd.Series(0, index=idx)).fillna(0)
    t3_pass = t1 & ~((ss == 1) | (be == 1) | (es == 1))

    hc_col = df.get("hardcheck_pass", pd.Series(1, index=idx)).astype(float).fillna(1)
    hc_pass = hc_col.astype(bool)

    passed = t1 & t2_pass & t3_pass & hc_pass

    # Build per-row fail reasons (for diagnostics)
    reasons = pd.Series("", index=idx)
    reasons[~t1] = reasons[~t1] + "T1:prob_low "
    reasons[t1 & ~t2_pass] = reasons[t1 & ~t2_pass] + "T2:vol_low "
    reasons[t1 & ~t3_pass] = reasons[t1 & ~t3_pass] + "T3:bearish_candle "
    reasons[t1 & ~hc_pass] = reasons[t1 & ~hc_pass] + "HC:fail "

    return passed, reasons


# ── sector cutoff filter ───────────────────────────────────────────────────────

def apply_sector_cutoffs(df: pd.DataFrame, prob: np.ndarray,
                         sector_map: dict, sector_cutoffs: dict,
                         base_cutoff: float = BASE_CUTOFF) -> pd.Series:
    """Returns bool mask — True = passes sector-specific cutoff."""
    result = pd.Series(True, index=df.index)
    for i, row in df.iterrows():
        sym = row.get("symbol", "")
        sec = sector_map.get(sym, "Unknown")
        thr = sector_cutoffs.get(sec, base_cutoff)
        if prob[df.index.get_loc(i)] < thr:
            result[i] = False
    return result


# ── main ───────────────────────────────────────────────────────────────────────

def run():
    # ── 1. Load training set (ALL rows — NO label_valid filter) ──────────────
    print(f"Loading training set: {TRAIN_SET.name} …")
    raw = pd.read_parquet(TRAIN_SET)
    raw = raw.reset_index()
    raw["date"] = pd.to_datetime(raw["date"])
    train_max = raw["date"].max()
    print(f"  Training set max date: {train_max.date()}")

    # Filter to our date window
    data_train = raw[(raw["date"] >= DATE_START) & (raw["date"] <= DATE_END)].copy()
    print(f"  From training set ({DATE_START}→{DATE_END}): {len(data_train):,} rows "
          f"({data_train['date'].min().date()} → {data_train['date'].max().date()})")

    # ── 2. Supplement: feature files for dates AFTER training set ─────────────
    # For dates > train_max, load directly from data/features/<date>.parquet
    supplement_rows = []
    extra_start = (train_max + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
    feat_files = sorted(FEAT_DIR.glob("*.parquet"))
    extra_dates = [
        f for f in feat_files
        if extra_start <= f.stem <= DATE_END
    ]
    if extra_dates:
        print(f"  Supplementing {len(extra_dates)} feature file(s): "
              f"{extra_dates[0].stem} → {extra_dates[-1].stem}")
        model_cols_needed = None
        for feat_file in extra_dates:
            try:
                fdf = pd.read_parquet(feat_file)
                fdf["date"] = pd.to_datetime(feat_file.stem)
                # Reset index if symbol is index
                if fdf.index.name == "symbol" or "symbol" not in fdf.columns:
                    fdf = fdf.reset_index()
                    if fdf.columns[0] != "symbol":
                        fdf = fdf.rename(columns={fdf.columns[0]: "symbol"})
                fdf["label_valid"] = 0   # no T+20 outcome yet
                fdf["label"]       = np.nan
                supplement_rows.append(fdf)
            except Exception as e:
                print(f"    ⚠️ Skip {feat_file.stem}: {e}")

    # Merge training + supplement
    if supplement_rows:
        sup_df = pd.concat(supplement_rows, ignore_index=True)
        print(f"  Supplement rows: {len(sup_df):,}")
        data = pd.concat([data_train, sup_df], ignore_index=True)
    else:
        data = data_train.copy()

    data = data.reset_index(drop=True)
    print(f"  Combined total: {len(data):,} rows  "
          f"({data['date'].min().date()} → {data['date'].max().date()})")

    # ── 2. Score ──────────────────────────────────────────────────────────────
    print("Scoring model 20D …")
    model, feat_cols = load_model()
    tmp = data.copy()
    for c in feat_cols:
        if c not in tmp.columns:
            tmp[c] = np.nan
    prob = model.predict_proba(tmp[feat_cols])[:, 1]
    data["prob"] = prob
    print(f"  Scored {len(data):,} rows  |  prob >= {BASE_CUTOFF}: {(prob >= BASE_CUTOFF).sum()}")

    # ── 3. Regime detection ───────────────────────────────────────────────────
    print("Detecting regimes …")
    from pipeline.regime_detector import detect_regime, REGIME_CUTOFFS
    unique_dates = sorted(data["date"].dt.strftime("%Y-%m-%d").unique())
    regime_map   = {}
    cutoff_map   = {}
    for d in unique_dates:
        try:
            r, dyn_cutoff, _ = detect_regime(d)
        except Exception:
            r, dyn_cutoff = "unknown", BASE_CUTOFF
        regime_map[d]  = r
        cutoff_map[d]  = dyn_cutoff
    data["regime"]        = data["date"].dt.strftime("%Y-%m-%d").map(regime_map)
    data["dynamic_cutoff"] = data["date"].dt.strftime("%Y-%m-%d").map(cutoff_map)
    print("  Regime distribution:")
    rc = data.groupby("regime")["date"].nunique().sort_values(ascending=False)
    for r, n in rc.items():
        print(f"    {r:<12} {n:>3} ngày")

    # ── 4. Sector map ─────────────────────────────────────────────────────────
    print("Loading sector map …")
    all_syms = data["symbol"].unique().tolist()
    from pipeline.backtest_live_system import load_sector_map
    sector_map = load_sector_map(all_syms)
    data["sector"] = data["symbol"].map(lambda s: sector_map.get(s, "Unknown"))

    # ── 5. Apply live filters per-row (dynamic cutoff + sector override) ───────
    print("Applying live filters (T1 base + T2 vol + T3 candle + HC + sector cutoffs) …")

    from pipeline.phase4b_ml_model_20d import SECTOR_CUTOFFS, SECTOR_CUTOFFS_DEEP_BEAR

    filter_rows = []
    for _, row in data.iterrows():
        regime   = row["regime"]
        sym      = row["symbol"]
        sec      = sector_map.get(sym, "Unknown")
        p        = float(row["prob"])
        date_str = row["date"].strftime("%Y-%m-%d")

        # Dynamic base cutoff (regime-aware)
        dyn_cut = float(row["dynamic_cutoff"])

        # Sector cutoff (regime-aware)
        sc_dict = SECTOR_CUTOFFS_DEEP_BEAR if regime == "deep_bear" else SECTOR_CUTOFFS
        eff_cut = sc_dict.get(sec, dyn_cut)

        # T1: prob
        t1_pass = (p >= eff_cut)

        # T2: vol_ratio
        vol = row.get("vol_ratio_20d", np.nan)
        t2_pass = pd.isna(vol) or (float(vol) >= 0.70)

        # T3: bearish candles
        ss = row.get("candle_shooting_star",     0) or 0
        be = row.get("candle_bearish_engulfing", 0) or 0
        es = row.get("candle_evening_star",      0) or 0
        t3_pass = not (ss == 1 or be == 1 or es == 1)

        # HC
        hc = row.get("hardcheck_pass", 1)
        hc_pass = bool(float(hc) == 1) if not pd.isna(hc) else True

        # Recovery skip
        skip_recovery = (regime == "recovery")

        all_pass = t1_pass and t2_pass and t3_pass and hc_pass and not skip_recovery

        # Build fail reason
        reasons = []
        if not t1_pass:    reasons.append(f"T1:prob({p:.3f}<{eff_cut:.2f})")
        if not t2_pass:    reasons.append(f"T2:vol({float(vol):.2f}<0.70)")
        if not t3_pass:    reasons.append("T3:bearish_candle")
        if not hc_pass:    reasons.append("HC:fail")
        if skip_recovery:  reasons.append("skip:recovery")

        filter_rows.append({
            "pass":         all_pass,
            "eff_cutoff":   eff_cut,
            "fail_reason":  " | ".join(reasons) if reasons else "",
            "t1_pass":      t1_pass,
            "t2_pass":      t2_pass,
            "t3_pass":      t3_pass,
            "hc_pass":      hc_pass,
        })

    fdf = pd.DataFrame(filter_rows, index=data.index)
    data = pd.concat([data, fdf], axis=1)

    passed = data[data["pass"]].copy()
    print(f"  Passed all filters: {len(passed):,} signals  "
          f"(from {(data['prob'] >= BASE_CUTOFF).sum()} at base cutoff)")

    # Breakdown
    all_above = data[data["prob"] >= BASE_CUTOFF]
    t1f  = (~fdf["t1_pass"]).sum()
    t2f  = (fdf["t1_pass"] & ~fdf["t2_pass"]).sum()
    t3f  = (fdf["t1_pass"] & ~fdf["t3_pass"]).sum()
    hcf  = (fdf["t1_pass"] & ~fdf["hc_pass"]).sum()
    recf = (data[data["pass"] == False]["regime"] == "recovery").sum()
    print(f"  Filter breakdown:")
    print(f"    T1 fail (sector/dynamic cutoff)  : {t1f}")
    print(f"    T2 fail (vol_ratio < 0.70)        : {t2f}")
    print(f"    T3 fail (bearish candle)          : {t3f}")
    print(f"    HC fail (hardcheck_pass=0)        : {hcf}")
    print(f"    Skipped (recovery regime)         : {recf}")

    # ── 6. FIRST_ONLY dedup ────────────────────────────────────────────────────
    print("Applying FIRST_ONLY dedup (20d cooldown) …")
    passed = passed.sort_values(["date", "prob"], ascending=[True, False])
    stock_last: dict = {}
    dedup_rows = []
    for _, row in passed.iterrows():
        sym, date = row["symbol"], row["date"]
        last = stock_last.get(sym)
        if last is not None and (date - last).days <= HOLD_DAYS:
            continue
        stock_last[sym] = date
        dedup_rows.append(row)
    eligible = pd.DataFrame(dedup_rows).reset_index(drop=True)
    print(f"  After FIRST_ONLY: {len(eligible):,} signals")

    # ── 7. Load prices & compute P&L ──────────────────────────────────────────
    print("Loading OHLCV prices …")
    pm  = load_price_map()
    cal = get_calendar(pm)
    print(f"  {len(pm)} symbols  |  {len(cal)} trading days")

    today = pd.Timestamp.today().normalize()

    rows_out = []
    for _, row in eligible.iterrows():
        sym       = row["symbol"]
        sec       = row.get("sector", "Unknown")
        regime    = row["regime"]
        ed        = row["date"]
        p         = round(float(row["prob"]), 4)
        eff_cut   = round(float(row["eff_cutoff"]), 2)
        dyn_cut   = round(float(row["dynamic_cutoff"]), 2)

        xd        = exit_date(ed, cal, HOLD_DAYS)   # planned T+20 exit date
        ep        = get_price(sym, ed, pm)           # entry price (close on signal date)

        # Determine if T+20 has arrived yet
        t20_arrived = (xd is not None) and (xd <= today)

        if t20_arrived:
            xp = get_price(sym, xd, pm)
        else:
            xp = None  # still open — no official exit yet

        # Current price (latest available) for open positions
        cur_p = get_price(sym, today, pm) if not t20_arrived else None

        if ep and t20_arrived and xp:
            # CLOSED: T+20 exit price known
            pnl_pct = round((xp - ep) / ep * 100, 2)
            pnl_vnd = int(CAPITAL_PER * pnl_pct / 100)
            cur_pct = pnl_pct   # same as closed pct
            if pnl_pct >= WIN_THRESH:
                outcome, label = "✅ WIN (+5%)", "WIN"
            elif pnl_pct >= 0:
                outcome, label = "⚠️ Hòa vốn", "HÒA"
            elif pnl_pct >= SL_THRESH:
                outcome, label = "❌ Lỗ", "LỖ"
            else:
                outcome, label = "🛑 SL (-10%+)", "SL"
            status = "Đã đóng"
        elif ep and not t20_arrived:
            # OPEN: show current price performance
            xp = None
            pnl_vnd = None
            if cur_p:
                pnl_pct = round((cur_p - ep) / ep * 100, 2)
                cur_pct = pnl_pct
            else:
                pnl_pct = None
                cur_pct = None
            outcome, label, status = "🔄 Đang mở", "MỞ", "Đang mở"
        else:
            xp = None; pnl_pct = None; pnl_vnd = None; cur_pct = None; cur_p = None
            outcome, label, status = "—", "—", "Không có giá"

        rows_out.append({
            "Tháng":              ed.strftime("%Y-%m"),
            "Mã CP":              sym,
            "Sector":             sec,
            "Regime":             regime,
            "Ngày vào":           ed.strftime("%d/%m/%Y"),
            "Ngày thoát T+20":    xd.strftime("%d/%m/%Y") if xd else "—",
            "Giá vào (VND)":      int(ep)   if ep   else None,
            "Giá thoát (VND)":    int(xp)   if xp   else None,
            "Giá hiện tại (VND)": int(cur_p) if cur_p else None,
            "% T+20":             pnl_pct if status == "Đã đóng" else None,
            "% Hiện tại":         cur_pct,
            "P&L ước tính (VND)": int(CAPITAL_PER * cur_pct / 100) if cur_pct is not None else None,
            "Kết quả":            outcome,
            "Label":              label,
            "Status":             status,
            "Prob":               p,
            "Dynamic Cutoff":     dyn_cut,
            "Eff Cutoff":         eff_cut,
        })

    result_df = pd.DataFrame(rows_out)
    closed_df = result_df[result_df["Status"] == "Đã đóng"].copy()
    open_df   = result_df[result_df["Status"] == "Đang mở"].copy()

    # ── 8. Stats ───────────────────────────────────────────────────────────────
    def _stats(sub, with_pnl=True):
        n = len(sub)
        if n == 0:
            return dict(n=0, wr=0.0, wr0=0.0, avg=0.0, pnl=0, sharpe=0.0)
        pcts = sub["% T+20"].dropna()
        win  = (pcts >= WIN_THRESH).sum()
        pos  = (pcts >= 0).sum()
        avg  = pcts.mean() if len(pcts) else 0
        std  = pcts.std()  if len(pcts) > 1 else 1
        sh   = avg / std * (252 / HOLD_DAYS) ** 0.5 if std > 0 else 0
        pnl  = int(sub["P&L ước tính (VND)"].dropna().sum()) if with_pnl else 0
        return dict(
            n=n, wr=round(win/n*100, 1), wr0=round(pos/n*100, 1),
            avg=round(avg, 2), pnl=pnl, sharpe=round(sh, 2)
        )

    total  = _stats(closed_df)
    by_sec = closed_df.groupby("Sector", group_keys=False).apply(_stats, include_groups=False).to_dict()
    by_reg = closed_df.groupby("Regime", group_keys=False).apply(_stats, include_groups=False).to_dict()

    # Open positions: unrealized P&L from current price
    open_cur_pct = open_df["% Hiện tại"].dropna()
    open_pnl_est = open_df["P&L ước tính (VND)"].dropna().sum()

    # ── 9. Console output ──────────────────────────────────────────────────────
    W = 72
    print("\n" + "=" * W)
    print(f"  SIGNAL DETAIL  {DATE_START} → {DATE_END}  |  ALL ELIGIBLE (no slot cap)")
    print("=" * W)
    print(f"  Eligible (sau filter + dedup): {len(eligible):,}")
    print(f"  Đã đóng (có kết quả T+20):    {total['n']:,}")
    print(f"  Đang mở (chưa đến T+20):      {len(open_df):,}")
    print(f"\n  ── Đã đóng ──")
    print(f"  WR ≥+5%  : {total['wr']:.1f}%")
    print(f"  WR ≥ 0%  : {total['wr0']:.1f}%")
    print(f"  Avg Ret  : {total['avg']:+.2f}%")
    print(f"  Sharpe   : {total['sharpe']:.2f}")
    print(f"  P&L thực : {total['pnl']/1e6:+.1f}M VND  (100M/lệnh)")
    if len(open_df):
        print(f"\n  ── Đang mở ({len(open_df)} lệnh) ──")
        print(f"  P&L ước tính (giá hôm nay): {open_pnl_est/1e6:+.1f}M VND")
        if len(open_cur_pct):
            print(f"  Avg % hiện tại: {open_cur_pct.mean():+.2f}%")

    print("\n── Theo Sector (đã đóng) ──")
    hdr = f"  {'Sector':<28}  {'n':>4}  {'WR%':>5}  {'WR0%':>5}  {'Avg':>7}  {'P&L(M)':>8}"
    print(hdr); print("  " + "-" * 64)
    for sec, s in sorted(by_sec.items(), key=lambda x: -x[1]["n"]):
        pnl_m = s["pnl"] / 1e6
        print(f"  {sec:<28}  {s['n']:>4}  {s['wr']:>5.1f}  {s['wr0']:>5.1f}  "
              f"{s['avg']:>+7.2f}  {pnl_m:>+8.1f}M")

    print("\n── Theo Regime (đã đóng) ──")
    hdr2 = f"  {'Regime':<14}  {'n':>4}  {'WR%':>5}  {'WR0%':>5}  {'Avg':>7}"
    print(hdr2); print("  " + "-" * 44)
    for reg, s in sorted(by_reg.items(), key=lambda x: -x[1]["n"]):
        print(f"  {reg:<14}  {s['n']:>4}  {s['wr']:>5.1f}  {s['wr0']:>5.1f}  {s['avg']:>+7.2f}")

    print("\n── Chi tiết tất cả signal ──")
    display_cols = ["Tháng","Mã CP","Sector","Regime","Ngày vào",
                    "Ngày thoát T+20","Giá vào (VND)","Giá thoát (VND)",
                    "Giá hiện tại (VND)","% T+20","% Hiện tại","Kết quả","Prob","Eff Cutoff"]
    print(result_df[display_cols].to_string(index=False))

    # ── 10. Write Excel ────────────────────────────────────────────────────────
    print(f"\nWriting Excel → {OUTPUT_PATH.name} …")

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:

        # ── Sheet: Tổng kết ─────────────────────────────────────────────────
        sum_rows = [
            ["AgentQ 20D — Signal Detail (No Slot Cap)", ""],
            ["Khoảng thời gian", f"{DATE_START} → {DATE_END}"],
            ["Model", "v3a LightGBM 20D"],
            ["Base Cutoff", BASE_CUTOFF],
            ["Hold Period", f"{HOLD_DAYS} ngày giao dịch"],
            ["Capital/lệnh", f"{CAPITAL_PER:,.0f} VND (100M)"],
            ["", ""],
            ["── ELIGIBLE SIGNALS ──", ""],
            ["Tổng eligible (sau filter + dedup)", len(eligible)],
            ["  Đã đóng (có kết quả T+20)", total["n"]],
            ["  Đang mở (chưa đến T+20)", len(open_df)],
            ["", ""],
            ["── HIỆU SUẤT ĐÃ ĐÓNG ──", ""],
            ["Win Rate ≥+5%",      f"{total['wr']:.1f}%"],
            ["Win Rate ≥ 0%",      f"{total['wr0']:.1f}%"],
            ["Avg Return/lệnh",    f"{total['avg']:+.2f}%"],
            ["Sharpe Ratio",       f"{total['sharpe']:.2f}"],
            ["P&L Thực (100M/lệnh)", f"{total['pnl']/1e6:+.1f} triệu VND"],
            ["", ""],
            ["── ĐANG MỞ (giá hôm nay) ──", ""],
            ["Số lệnh đang mở",    len(open_df)],
            ["P&L ước tính",       f"{open_pnl_est/1e6:+.1f} triệu VND"],
            ["Avg % hiện tại",     f"{open_cur_pct.mean():+.2f}%" if len(open_cur_pct) else "—"],
            ["", ""],
            ["── GHI CHÚ ──", ""],
            ["WIN",  "Lãi ≥ +5%  (model target)"],
            ["HÒA",  "Lãi  0 ~ +5%"],
            ["LỖ",   "Lỗ   0 ~ -10%"],
            ["SL",   "Lỗ  > -10%  (stop-loss triggered)"],
        ]
        pd.DataFrame(sum_rows, columns=["Chỉ tiêu", "Giá trị"]).to_excel(
            writer, sheet_name="📊 Tổng kết", index=False)

        # Sector breakdown
        sec_rows = []
        for sec, s in sorted(by_sec.items(), key=lambda x: -x[1]["n"]):
            sec_rows.append({
                "Sector": sec, "n": s["n"],
                "WR ≥+5%": f"{s['wr']:.1f}%",
                "WR ≥ 0%": f"{s['wr0']:.1f}%",
                "Avg Ret": s["avg"],
                "P&L (VND)": s["pnl"],
            })
        if sec_rows:
            startrow = len(sum_rows) + 3
            pd.DataFrame([["── THEO SECTOR ──"]], columns=["Sector"]).to_excel(
                writer, sheet_name="📊 Tổng kết", index=False, startrow=startrow - 1, header=False)
            pd.DataFrame(sec_rows).to_excel(
                writer, sheet_name="📊 Tổng kết", index=False, startrow=startrow)

        # Regime breakdown
        reg_rows = []
        for reg, s in sorted(by_reg.items(), key=lambda x: -x[1]["n"]):
            reg_rows.append({
                "Regime": reg, "n": s["n"],
                "WR ≥+5%": f"{s['wr']:.1f}%",
                "WR ≥ 0%": f"{s['wr0']:.1f}%",
                "Avg Ret": s["avg"],
            })
        if reg_rows:
            startrow2 = len(sum_rows) + 3 + len(sec_rows) + 5
            pd.DataFrame([["── THEO REGIME ──"]], columns=["Regime"]).to_excel(
                writer, sheet_name="📊 Tổng kết", index=False, startrow=startrow2 - 1, header=False)
            pd.DataFrame(reg_rows).to_excel(
                writer, sheet_name="📊 Tổng kết", index=False, startrow=startrow2)

        # ── Sheet: All Signal ───────────────────────────────────────────────
        all_cols = ["Tháng","Mã CP","Sector","Regime","Ngày vào",
                    "Ngày thoát T+20","Giá vào (VND)","Giá thoát (VND)",
                    "Giá hiện tại (VND)","% T+20","% Hiện tại",
                    "P&L ước tính (VND)","Kết quả","Label","Status",
                    "Prob","Dynamic Cutoff","Eff Cutoff"]
        result_df[all_cols].to_excel(
            writer, sheet_name="📋 All Signal (No Cap)", index=False)

        # ── Sheet: Closed only ──────────────────────────────────────────────
        closed_df[all_cols].sort_values("Ngày vào").to_excel(
            writer, sheet_name="✅ Đã đóng", index=False)

        # ── Sheet: Open ─────────────────────────────────────────────────────
        if len(open_df):
            open_df[all_cols].sort_values("Ngày vào").to_excel(
                writer, sheet_name="🔄 Đang mở", index=False)

        # ── Sheet: Monthly ──────────────────────────────────────────────────
        for mkey in sorted(result_df["Tháng"].unique()):
            mdf = result_df[result_df["Tháng"] == mkey][all_cols]
            if len(mdf):
                month_label = f"T{mkey[5:7]}-{mkey[:4]}"
                mdf.to_excel(writer, sheet_name=month_label, index=False)

        # ── Formatting ───────────────────────────────────────────────────────
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        FMT = {
            "WIN":  PatternFill("solid", fgColor="C6EFCE"),
            "HÒA":  PatternFill("solid", fgColor="FFEB9C"),
            "LỖ":   PatternFill("solid", fgColor="FFC7CE"),
            "SL":   PatternFill("solid", fgColor="FF0000"),
            "MỞ":   PatternFill("solid", fgColor="BDD7EE"),
            "HDR":  PatternFill("solid", fgColor="1F4E79"),
        }
        F_WHITE = Font(color="FFFFFF", bold=True)
        SL_FONT = Font(color="FFFFFF", bold=True)

        for ws in writer.sheets.values():
            for col in ws.columns:
                ml = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 35)
            for cell in ws[1]:
                if cell.value:
                    cell.fill = FMT["HDR"]
                    cell.font = F_WHITE
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            header = [c.value for c in ws[1]]
            lbl_col = header.index("Label") + 1 if "Label" in header else None
            pct_col = header.index("% Lãi/Lỗ") + 1 if "% Lãi/Lỗ" in header else None
            for row in ws.iter_rows(min_row=2):
                lbl = row[lbl_col - 1].value if lbl_col else None
                fill = FMT.get(str(lbl), None)
                if fill:
                    for cell in row:
                        cell.fill = fill
                    if str(lbl) == "SL":
                        for cell in row:
                            cell.font = SL_FONT
                if pct_col:
                    pct_cell = row[pct_col - 1]
                    if isinstance(pct_cell.value, float):
                        pct_cell.number_format = "+0.00;-0.00"

    # ── Copy to Desktop ────────────────────────────────────────────────────────
    import shutil
    dest = Path.home() / "Desktop" / "signal_detail_2026_Mar_May_noslot.xlsx"
    shutil.copy(OUTPUT_PATH, dest)
    print(f"\n✅ Done!")
    print(f"   Saved → {dest}")


if __name__ == "__main__":
    run()
