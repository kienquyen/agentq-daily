#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline/export_signals_excel.py

Export 2026 signal list to Excel with:
  - Entry/exit dates (T+20 trading days)
  - Entry/exit prices (close VND)
  - % profit/loss
  - 100M VND simulation per trade
  - Monthly breakdown sheets + Summary sheet

Strategy:
  - v3a model @ cutoff 0.52
  - FIRST_ONLY: bỏ trùng cùng mã trong 20d window
  - Hiển thị TẤT CẢ eligible signals, đánh dấu lệnh nào được chọn
    khi giới hạn 15 slot đồng thời

Usage:
    python -m pipeline.export_signals_excel
"""
from __future__ import annotations
import sys, pickle
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

MODEL_PATH   = ROOT / "data/models/lgbm_model_20d.pkl"
TRAIN_SET    = ROOT / "data/labels/training_set_20d.parquet"
OHLCV_DIR    = ROOT / "data/ohlcv"
OUTPUT_PATH  = ROOT / "data/exports/signals_2026.xlsx"
CUTOFF       = 0.52
MAX_SLOTS    = 15
HOLD_DAYS    = 20        # trading days
CAPITAL_PER  = 100_000_000   # 100 triệu VND

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)


# ─── price helpers ────────────────────────────────────────────────────────────

def load_price_map() -> dict[str, pd.DataFrame]:
    price_map = {}
    for f in OHLCV_DIR.glob("*.parquet"):
        try:
            df = pd.read_parquet(f)[["time", "open", "close"]]
            df["time"] = pd.to_datetime(df["time"])
            price_map[f.stem] = df.set_index("time").sort_index()
        except Exception:
            pass
    return price_map


def get_trading_calendar(price_map: dict) -> list[pd.Timestamp]:
    all_dates: set = set()
    for df in price_map.values():
        all_dates.update(df.index.tolist())
    return sorted(all_dates)


def get_exit_date(entry: pd.Timestamp, cal: list, hold: int) -> pd.Timestamp | None:
    try:
        idx = cal.index(entry)
        ei = idx + hold
        return cal[ei] if ei < len(cal) else None
    except (ValueError, IndexError):
        return None


def get_price(sym: str, date: pd.Timestamp | None, price_map: dict) -> float | None:
    if date is None or sym not in price_map:
        return None
    df = price_map[sym]
    if date in df.index:
        return float(df.loc[date, "close"])
    for d in [1, -1, 2, -2]:
        alt = date + pd.Timedelta(days=d)
        if alt in df.index:
            return float(df.loc[alt, "close"])
    return None


# ─── model ────────────────────────────────────────────────────────────────────

def load_model():
    obj = pickle.load(open(MODEL_PATH, "rb"))
    return (obj["model"], obj["feat_cols"]) if isinstance(obj, dict) else (obj, [])


def score(model, df, feat_cols):
    tmp = df.copy()
    for c in feat_cols:
        if c not in tmp.columns:
            tmp[c] = np.nan
    return model.predict_proba(tmp[feat_cols])[:, 1]


# ─── signal filtering ─────────────────────────────────────────────────────────

def get_eligible_signals(signals_df: pd.DataFrame) -> pd.DataFrame:
    """FIRST_ONLY: one signal per stock per 20d cooldown window."""
    stock_last: dict[str, pd.Timestamp] = {}
    eligible = []
    for _, row in signals_df.sort_values(["date", "prob"], ascending=[True, False]).iterrows():
        sym, date = row["symbol"], row["date"]
        last = stock_last.get(sym)
        if last is not None and (date - last).days <= HOLD_DAYS:
            continue
        stock_last[sym] = date
        eligible.append(row)
    return pd.DataFrame(eligible).reset_index(drop=True)


def simulate_portfolio(eligible_df: pd.DataFrame, cal: list) -> set[tuple]:
    """Apply MAX_SLOTS cap. Returns set of (symbol, date) taken."""
    date_to_idx = {d: i for i, d in enumerate(cal)}
    open_pos: dict[str, int] = {}   # symbol → entry_cal_idx
    open_count = 0
    taken: set[tuple] = set()

    for date in sorted(eligible_df["date"].unique()):
        cur_idx = date_to_idx.get(date, -1)
        # expire positions
        to_close = [s for s, ei in open_pos.items() if (ei + HOLD_DAYS) <= cur_idx]
        for s in to_close:
            del open_pos[s]
            open_count -= 1
        # take new signals (highest prob first)
        day_sigs = eligible_df[eligible_df["date"] == date].sort_values("prob", ascending=False)
        for _, row in day_sigs.iterrows():
            sym = row["symbol"]
            if sym in open_pos or open_count >= MAX_SLOTS:
                continue
            open_pos[sym] = cur_idx
            open_count += 1
            taken.add((sym, date))

    return taken


# ─── main ─────────────────────────────────────────────────────────────────────

def run():
    # ── 1. Load & score ───────────────────────────────────────────────────────
    print("Loading data …")
    df = pd.read_parquet(TRAIN_SET)
    if "label_valid" in df.columns:
        df = df[df["label_valid"] == 1]
    df = df.reset_index()
    df["date"]     = pd.to_datetime(df["date"])
    df["ret_pct"]  = df["forward_ret_10d"] * 100

    data_2026 = df[df["date"] >= "2026-01-01"].copy()
    print(f"2026 rows: {len(data_2026):,}  "
          f"|  {data_2026['date'].min().date()} → {data_2026['date'].max().date()}")

    print("Scoring …")
    model, feat_cols = load_model()
    data_2026["prob"] = score(model, data_2026, feat_cols)

    # ── 2. Filter ─────────────────────────────────────────────────────────────
    raw_signals = (
        data_2026[data_2026["prob"] >= CUTOFF][["symbol", "date", "prob", "ret_pct"]]
        .copy()
        .sort_values(["date", "prob"], ascending=[True, False])
        .reset_index(drop=True)
    )
    print(f"Raw signals @ {CUTOFF}: {len(raw_signals):,}")

    eligible = get_eligible_signals(raw_signals)
    print(f"Eligible after FIRST_ONLY dedup: {len(eligible):,}")

    # ── 3. Prices ─────────────────────────────────────────────────────────────
    print("Loading OHLCV …")
    price_map = load_price_map()
    calendar  = get_trading_calendar(price_map)
    print(f"Loaded {len(price_map)} symbols | {len(calendar)} trading days")

    # ── 4. Portfolio simulation ───────────────────────────────────────────────
    taken_set = simulate_portfolio(eligible, calendar)
    print(f"Taken by portfolio (≤{MAX_SLOTS} slots): {len(taken_set)}")

    # ── 5. Build rows for ALL eligible (mark taken vs skipped) ────────────────
    rows = []
    for _, row in eligible.iterrows():
        sym        = row["symbol"]
        entry_date = row["date"]
        prob       = round(float(row["prob"]), 4)
        is_taken   = (sym, entry_date) in taken_set

        exit_date   = get_exit_date(entry_date, calendar, HOLD_DAYS)
        entry_price = get_price(sym, entry_date, price_map)
        exit_price  = get_price(sym, exit_date,  price_map)

        if entry_price and exit_price:
            pnl_pct = round((exit_price - entry_price) / entry_price * 100, 2)
            pnl_vnd = int(CAPITAL_PER * pnl_pct / 100) if is_taken else None
            outcome = "✅ WIN" if pnl_pct >= 5 else ("⚠️ Hòa" if pnl_pct >= 0 else "❌ Lỗ")
            status  = "Đã đóng"
        elif entry_price:
            pnl_pct = None
            pnl_vnd = None
            outcome = "🔄 Đang mở"
            status  = "Đang mở"
        else:
            pnl_pct = None
            pnl_vnd = None
            outcome = "—"
            status  = "Không có giá"

        rows.append({
            "Tháng":              entry_date.strftime("%Y-%m"),
            "Mã CP":              sym,
            "Ngày vào":           entry_date.strftime("%d/%m/%Y"),
            "Ngày thoát (T+20)":  exit_date.strftime("%d/%m/%Y") if exit_date else "—",
            "Giá vào (VND)":      int(entry_price) if entry_price else None,
            "Giá thoát (VND)":    int(exit_price)  if exit_price  else None,
            "% Lãi/Lỗ":          pnl_pct,
            "Lãi/Lỗ (VND)":      pnl_vnd,
            "Kết quả":            outcome,
            "Prob":               prob,
            "Chọn (15 slot)":     "✔ Chọn" if is_taken else "— Hết slot",
            "Trạng thái":         status,
        })

    result_df = pd.DataFrame(rows)

    # ── 6. Stats ──────────────────────────────────────────────────────────────
    taken_df    = result_df[result_df["Chọn (15 slot)"] == "✔ Chọn"].copy()
    skipped_df  = result_df[result_df["Chọn (15 slot)"] != "✔ Chọn"].copy()
    closed_tkn  = taken_df[taken_df["Trạng thái"] == "Đã đóng"]
    elig_closed = result_df[result_df["Trạng thái"] == "Đã đóng"]

    def _stats(sub):
        n = len(sub)
        if n == 0:
            return 0, 0, 0, 0, 0, 0
        win  = (sub["% Lãi/Lỗ"] >= 5).sum()
        be   = ((sub["% Lãi/Lỗ"] >= 0) & (sub["% Lãi/Lỗ"] < 5)).sum()
        loss = (sub["% Lãi/Lỗ"] < 0).sum()
        wr   = win / n * 100
        avg  = sub["% Lãi/Lỗ"].mean()
        pnl  = sub["Lãi/Lỗ (VND)"].sum() if "Lãi/Lỗ (VND)" in sub else 0
        return win, be, loss, wr, avg, pnl

    win_t, be_t, loss_t, wr_t, avg_t, pnl_t = _stats(closed_tkn)
    win_e, _,    _,      wr_e, avg_e, _     = _stats(elig_closed)

    # Monthly
    def monthly_stats(sub):
        return sub.groupby("Tháng").agg(
            Lệnh=("Mã CP", "count"),
            WIN=("% Lãi/Lỗ", lambda x: (x >= 5).sum()),
            Hòa=("% Lãi/Lỗ", lambda x: ((x >= 0) & (x < 5)).sum()),
            Lỗ=("% Lãi/Lỗ", lambda x: (x < 0).sum()),
            WR=("% Lãi/Lỗ", lambda x: round((x >= 5).mean() * 100, 1)),
            AvgRet=("% Lãi/Lỗ", lambda x: round(x.mean(), 2)),
        ).reset_index()

    mth_tkn  = monthly_stats(closed_tkn)
    mth_elig = monthly_stats(elig_closed)
    if len(closed_tkn):
        mth_tkn["PnL (Tr.VND)"] = closed_tkn.groupby("Tháng")["Lãi/Lỗ (VND)"].sum().values / 1e6

    # ── 7. Write Excel ────────────────────────────────────────────────────────
    print(f"Writing Excel → {OUTPUT_PATH}")

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:

        # ── Sheet 1: Tổng kết ─────────────────────────────────────────────────
        sum_rows = [
            ["AgentQ 20D — Báo cáo Signal 2026", ""],
            ["Model", "v3a (LightGBM + Foreign Flow)"],
            ["Cutoff xác suất", CUTOFF],
            ["Hold period", f"{HOLD_DAYS} ngày giao dịch"],
            ["Max vị thế đồng thời", MAX_SLOTS],
            ["Vốn mỗi lệnh", f"{CAPITAL_PER:,.0f} VND"],
            ["", ""],
            ["── LUỒNG LỌC ──", ""],
            ["① Raw signals (prob ≥ 0.52)", len(raw_signals)],
            ["② Eligible sau FIRST_ONLY dedup", len(eligible)],
            ["   Trong đó: được portfolio chọn (≤15 slot)", len(taken_set)],
            ["   Trong đó: bị bỏ qua (hết slot)", len(eligible) - len(taken_set)],
            ["", ""],
            ["── PORTFOLIO THỰC TẾ (15 slot) ──", ""],
            ["Tổng lệnh thực thi", len(taken_df)],
            ["  Đã đóng", len(closed_tkn)],
            ["  Đang mở", len(taken_df[taken_df["Trạng thái"] == "Đang mở"])],
            ["WIN (≥+5%)", int(win_t)],
            ["Hòa vốn (0~+5%)", int(be_t)],
            ["Lỗ (<0%)", int(loss_t)],
            ["Win Rate", f"{wr_t:.1f}%"],
            ["Avg Return/lệnh", f"{avg_t:+.2f}%"],
            ["Tổng PnL (đã đóng)", f"{pnl_t/1e6:+.1f} triệu VND"],
            ["", ""],
            ["── NẾU KHÔNG GIỚI HẠN SLOT (tất cả eligible) ──", ""],
            ["Tổng eligible đã đóng", len(elig_closed)],
            ["Win Rate (no cap)", f"{wr_e:.1f}%"],
            ["Avg Return (no cap)", f"{avg_e:+.2f}%"],
        ]
        pd.DataFrame(sum_rows, columns=["Chỉ tiêu", "Giá trị"]).to_excel(
            writer, sheet_name="📊 Tổng kết", index=False)

        ws = writer.sheets["📊 Tổng kết"]
        base = len(sum_rows) + 2

        ws.cell(base, 1, "── THEO THÁNG — Portfolio 15 slot ──")
        mth_tkn.to_excel(writer, sheet_name="📊 Tổng kết", index=False, startrow=base)

        base2 = base + len(mth_tkn) + 3
        ws.cell(base2, 1, "── THEO THÁNG — Tất cả eligible (không giới hạn slot) ──")
        mth_elig.to_excel(writer, sheet_name="📊 Tổng kết", index=False, startrow=base2)

        # ── Sheet 2: Tất cả signal (eligible) ────────────────────────────────
        all_cols = ["Tháng","Mã CP","Ngày vào","Ngày thoát (T+20)",
                    "Giá vào (VND)","Giá thoát (VND)","% Lãi/Lỗ",
                    "Lãi/Lỗ (VND)","Kết quả","Prob","Chọn (15 slot)"]
        result_df[all_cols].to_excel(writer, sheet_name="📋 Tất cả signal", index=False)

        # ── Sheet 3: Portfolio 15 slot ────────────────────────────────────────
        port_cols = ["Tháng","Mã CP","Ngày vào","Ngày thoát (T+20)",
                     "Giá vào (VND)","Giá thoát (VND)","% Lãi/Lỗ",
                     "Lãi/Lỗ (VND)","Kết quả","Prob"]
        taken_df[port_cols].to_excel(writer, sheet_name="💼 Portfolio (15 slot)", index=False)

        # ── Monthly sheets (tất cả eligible) ─────────────────────────────────
        month_labels = {
            "2026-01": "01 Tháng 1",
            "2026-02": "02 Tháng 2",
            "2026-03": "03 Tháng 3",
            "2026-04": "04 Tháng 4",
            "2026-05": "05 Tháng 5",
        }
        for mkey, mlabel in month_labels.items():
            mdf = result_df[result_df["Tháng"] == mkey][all_cols].copy()
            if len(mdf):
                mdf.to_excel(writer, sheet_name=mlabel, index=False)

        # ── Formatting ────────────────────────────────────────────────────────
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        C_GREEN  = PatternFill("solid", fgColor="C6EFCE")
        C_RED    = PatternFill("solid", fgColor="FFC7CE")
        C_YELLOW = PatternFill("solid", fgColor="FFEB9C")
        C_BLUE   = PatternFill("solid", fgColor="BDD7EE")
        C_GREY   = PatternFill("solid", fgColor="E8E8E8")
        C_HEADER = PatternFill("solid", fgColor="1F4E79")
        F_WHITE  = Font(color="FFFFFF", bold=True)
        F_BOLD   = Font(bold=True)

        for sheet in writer.sheets.values():
            # Column widths
            for col in sheet.columns:
                ml = max((len(str(c.value or "")) for c in col), default=8)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 32)

            # Header
            for cell in sheet[1]:
                if cell.value:
                    cell.fill   = C_HEADER
                    cell.font   = F_WHITE
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Row coloring by outcome
            header = [c.value for c in sheet[1]]
            kq_col = header.index("Kết quả") + 1 if "Kết quả" in header else None
            sl_col = header.index("Chọn (15 slot)") + 1 if "Chọn (15 slot)" in header else None

            for row in sheet.iter_rows(min_row=2):
                outcome = row[kq_col - 1].value if kq_col else None
                chosen  = row[sl_col  - 1].value if sl_col  else None

                # Dim skipped rows
                if chosen and "Hết slot" in str(chosen):
                    for cell in row:
                        cell.fill = C_GREY
                    continue

                if outcome:
                    fill = (C_GREEN  if "✅" in str(outcome) else
                            C_RED    if "❌" in str(outcome) else
                            C_YELLOW if "⚠️" in str(outcome) else
                            C_BLUE   if "🔄" in str(outcome) else None)
                    if fill:
                        for cell in row:
                            cell.fill = fill

                # Number formats
                for cell in row:
                    if isinstance(cell.value, (int, float)) and cell.value is not None:
                        col_name = header[cell.column - 1] if cell.column <= len(header) else ""
                        if "VND" in str(col_name):
                            cell.number_format = "#,##0"
                        elif "%" in str(col_name):
                            cell.number_format = "+0.00;-0.00"

    print(f"\n✅ Done! → {OUTPUT_PATH}")

    # ── Console summary ───────────────────────────────────────────────────────
    print("\n" + "=" * 72)
    print("TẤT CẢ ELIGIBLE SIGNALS 2026")
    print("=" * 72)
    print(result_df[["Tháng","Mã CP","Ngày vào","Ngày thoát (T+20)",
                      "Giá vào (VND)","Giá thoát (VND)",
                      "% Lãi/Lỗ","Kết quả","Prob","Chọn (15 slot)"]].to_string(index=False))

    print("\n" + "=" * 72)
    print("PORTFOLIO 15 SLOT — theo tháng")
    print("=" * 72)
    print(mth_tkn.to_string(index=False))
    print(f"\n  Win Rate: {wr_t:.1f}%  |  Avg: {avg_t:+.2f}%  |  PnL: {pnl_t/1e6:+.1f}M VND")

    print("\n" + "=" * 72)
    print("TẤT CẢ ELIGIBLE (no slot cap) — theo tháng")
    print("=" * 72)
    print(mth_elig.to_string(index=False))
    print(f"\n  Win Rate: {wr_e:.1f}%  |  Avg: {avg_e:+.2f}%")


if __name__ == "__main__":
    run()
