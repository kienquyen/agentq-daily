#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pipeline/compare_hold_period.py

So sánh hiệu suất giữ T+15 / T+20 / T+25 / T+30 trên toàn bộ lịch sử.

Quy trình:
  1. Load & score tất cả signals đã qua live filters (giống live system)
  2. Với mỗi signal: tính exit price tại T+15, T+20, T+25, T+30 từ OHLCV
  3. So sánh WR ≥+5%, WR ≥0%, avg, Sharpe, P&L theo từng hold period
  4. Breakdown theo regime + sector
  5. Export Excel + in console table

Usage:
    python -m pipeline.compare_hold_period
"""
from __future__ import annotations

import pickle, sys
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

MODEL_PATH = ROOT / "data/models/lgbm_model_20d.pkl"
TRAIN_SET  = ROOT / "data/labels/training_set_20d.parquet"
OHLCV_DIR  = ROOT / "data/ohlcv"
OUTPUT_XLS = ROOT / "data/exports/compare_hold_period.xlsx"
OUTPUT_XLS.parent.mkdir(parents=True, exist_ok=True)

TEST_START  = "2022-07-01"   # đủ dữ liệu OHLCV để tính T+30
BASE_CUTOFF = 0.52
HOLD_PERIODS = [15, 20, 25, 30]
WIN_THRESH   = 5.0    # % target
CAPITAL_PER  = 100_000_000


# ── price helpers ──────────────────────────────────────────────────────────────

def load_price_map() -> dict[str, pd.DataFrame]:
    pm = {}
    for f in OHLCV_DIR.glob("*.parquet"):
        try:
            df = pd.read_parquet(f)[["time", "close"]]
            df["time"] = pd.to_datetime(df["time"])
            pm[f.stem] = df.set_index("time").sort_index()
        except Exception:
            pass
    return pm


def get_calendar(pm: dict) -> list[pd.Timestamp]:
    s: set = set()
    for df in pm.values():
        s.update(df.index.tolist())
    return sorted(s)


def get_price(sym: str, date, pm: dict, cal_set: set):
    """Return close price on or near date (±2 trading days)."""
    if date is None or sym not in pm:
        return None
    df = pm[sym]
    if date in df.index:
        return float(df.loc[date, "close"])
    for d in [1, -1, 2, -2]:
        alt = date + pd.Timedelta(days=d)
        if alt in df.index and alt in cal_set:
            return float(df.loc[alt, "close"])
    return None


def exit_date_at(entry: pd.Timestamp, cal: list, hold: int):
    try:
        idx = cal.index(entry)
        ei = idx + hold
        return cal[ei] if ei < len(cal) else None
    except (ValueError, IndexError):
        return None


# ── stats ──────────────────────────────────────────────────────────────────────

def _stats(pcts: pd.Series, capital=CAPITAL_PER) -> dict:
    pcts = pcts.dropna()
    n = len(pcts)
    if n == 0:
        return dict(n=0, wr5=0.0, wr0=0.0, avg=0.0, med=0.0,
                    sharpe=0.0, pnl_m=0.0, sl_pct=0.0)
    win5  = (pcts >= WIN_THRESH).sum()
    win0  = (pcts >= 0).sum()
    sl    = (pcts <= -10).sum()
    avg   = pcts.mean()
    med   = pcts.median()
    std   = pcts.std() if n > 1 else 1e-9
    ann   = (252 / 20) ** 0.5   # annualised using 20d base
    sh    = avg / std * ann if std > 0 else 0
    pnl_m = pcts.sum() * capital / 1e8   # in triệu VND (100M each)
    return dict(
        n=n,
        wr5=round(win5 / n * 100, 1),
        wr0=round(win0 / n * 100, 1),
        avg=round(avg, 2),
        med=round(med, 2),
        sharpe=round(sh, 2),
        pnl_m=round(pnl_m, 1),
        sl_pct=round(sl / n * 100, 1),
    )


# ── main ───────────────────────────────────────────────────────────────────────

def run():
    # ── 1. Load training set (label_valid=1 only — has forward returns) ───────
    print("Loading training set …")
    raw = pd.read_parquet(TRAIN_SET)
    if "label_valid" in raw.columns:
        raw = raw[raw["label_valid"] == 1]
    raw = raw.reset_index()
    raw["date"] = pd.to_datetime(raw["date"])
    data = raw[raw["date"] >= TEST_START].copy().reset_index(drop=True)
    print(f"  Rows: {len(data):,}  |  {data['date'].min().date()} → {data['date'].max().date()}")

    # ── 2. Score ──────────────────────────────────────────────────────────────
    print("Scoring model 20D …")
    obj = pickle.load(open(MODEL_PATH, "rb"))
    model, feat_cols = obj["model"], obj["feat_cols"]
    tmp = data.copy()
    for c in feat_cols:
        if c not in tmp.columns:
            tmp[c] = np.nan
    data["prob"] = model.predict_proba(tmp[feat_cols])[:, 1]

    # ── 3. Regime detection ───────────────────────────────────────────────────
    print("Detecting regimes …")
    from pipeline.regime_detector import detect_regime
    unique_dates = sorted(data["date"].dt.strftime("%Y-%m-%d").unique())
    regime_map, cutoff_map = {}, {}
    for d in unique_dates:
        try:
            r, c, _ = detect_regime(d)
        except Exception:
            r, c = "unknown", BASE_CUTOFF
        regime_map[d] = r
        cutoff_map[d] = c
    data["regime"]  = data["date"].dt.strftime("%Y-%m-%d").map(regime_map)
    data["dyn_cut"] = data["date"].dt.strftime("%Y-%m-%d").map(cutoff_map)

    # ── 4. Sector map ─────────────────────────────────────────────────────────
    print("Loading sector map …")
    from pipeline.backtest_live_system import load_sector_map
    sector_map = load_sector_map(data["symbol"].unique().tolist())
    data["sector"] = data["symbol"].map(lambda s: sector_map.get(s, "Unknown"))

    # ── 5. Live filters (per-row, regime+sector aware) ────────────────────────
    print("Applying live filters …")
    from pipeline.phase4b_ml_model_20d import SECTOR_CUTOFFS, SECTOR_CUTOFFS_DEEP_BEAR

    pass_mask = []
    eff_cuts  = []
    for _, row in data.iterrows():
        regime = row["regime"]
        sec    = sector_map.get(row["symbol"], "Unknown")
        p      = float(row["prob"])
        dyn    = float(row["dyn_cut"])
        sc     = SECTOR_CUTOFFS_DEEP_BEAR if regime == "deep_bear" else SECTOR_CUTOFFS
        eff    = sc.get(sec, dyn)
        eff_cuts.append(eff)

        t1 = p >= eff
        vol = row.get("vol_ratio_20d", np.nan)
        t2  = pd.isna(vol) or float(vol) >= 0.70
        ss  = row.get("candle_shooting_star",     0) or 0
        be  = row.get("candle_bearish_engulfing", 0) or 0
        es  = row.get("candle_evening_star",      0) or 0
        t3  = not (ss == 1 or be == 1 or es == 1)
        hc  = row.get("hardcheck_pass", 1)
        hc_pass = bool(float(hc) == 1) if not pd.isna(hc) else True
        skip_rec = regime == "recovery"

        pass_mask.append(t1 and t2 and t3 and hc_pass and not skip_rec)

    data["pass"]    = pass_mask
    data["eff_cut"] = eff_cuts
    passed = data[data["pass"]].copy()
    print(f"  Passed filters: {len(passed):,} signals")

    # ── 6. FIRST_ONLY dedup (20d window) ──────────────────────────────────────
    print("FIRST_ONLY dedup …")
    passed = passed.sort_values(["date", "prob"], ascending=[True, False])
    stock_last: dict = {}
    eligible_rows = []
    for _, row in passed.iterrows():
        sym, date = row["symbol"], row["date"]
        last = stock_last.get(sym)
        if last is not None and (date - last).days <= 20:
            continue
        stock_last[sym] = date
        eligible_rows.append(row)
    eligible = pd.DataFrame(eligible_rows).reset_index(drop=True)
    print(f"  After dedup: {len(eligible):,} signals")
    print(f"  Date range : {eligible['date'].min().date()} → {eligible['date'].max().date()}")

    # ── 7. Load OHLCV + calendar ──────────────────────────────────────────────
    print("Loading OHLCV prices …")
    pm      = load_price_map()
    cal     = get_calendar(pm)
    cal_set = set(cal)
    print(f"  {len(pm)} symbols  |  {len(cal)} trading days")

    # ── 8. Compute returns for each hold period ────────────────────────────────
    print(f"Computing returns for hold periods {HOLD_PERIODS} …")

    rows_out = []
    for _, row in eligible.iterrows():
        sym    = row["symbol"]
        ed     = row["date"]
        ep     = get_price(sym, ed, pm, cal_set)
        if ep is None:
            continue

        ret_by_hold = {}
        for h in HOLD_PERIODS:
            xd = exit_date_at(ed, cal, h)
            xp = get_price(sym, xd, pm, cal_set) if xd else None
            if xp:
                ret_by_hold[h] = round((xp - ep) / ep * 100, 2)
            else:
                ret_by_hold[h] = None

        # Only keep row if at least T+20 is available
        if ret_by_hold.get(20) is None:
            continue

        rows_out.append({
            "date":    ed,
            "symbol":  sym,
            "sector":  row["sector"],
            "regime":  row["regime"],
            "prob":    round(float(row["prob"]), 4),
            "eff_cut": round(float(row["eff_cut"]), 2),
            **{f"ret_t{h}": ret_by_hold[h] for h in HOLD_PERIODS},
        })

    df = pd.DataFrame(rows_out)
    print(f"  Signals with T+20 price: {len(df):,}")

    # ── 9. Global comparison ──────────────────────────────────────────────────
    W = 72
    print("\n" + "=" * W)
    print("  COMPARISON: HOLD PERIOD  T+15 / T+20 / T+25 / T+30")
    print(f"  {df['date'].min().date()} → {df['date'].max().date()}  |  n={len(df):,} signals")
    print("=" * W)

    global_stats = {}
    hdr = (f"  {'Hold':>5}  {'n':>5}  {'WR≥5%':>6}  {'WR≥0%':>6}  "
           f"{'Avg%':>6}  {'Med%':>6}  {'Sharpe':>7}  {'P&L(M)':>8}  {'SL%':>5}")
    print(hdr)
    print("  " + "-" * (len(hdr) - 2))
    for h in HOLD_PERIODS:
        col = f"ret_t{h}"
        s   = _stats(df[col])
        global_stats[h] = s
        print(f"  T+{h:<3}  {s['n']:>5}  {s['wr5']:>6.1f}  {s['wr0']:>6.1f}  "
              f"{s['avg']:>+6.2f}  {s['med']:>+6.2f}  {s['sharpe']:>7.2f}  "
              f"{s['pnl_m']:>+8.1f}  {s['sl_pct']:>5.1f}")

    # ── 10. By regime ─────────────────────────────────────────────────────────
    print(f"\n{'─'*W}")
    print("  THEO REGIME")
    print(f"{'─'*W}")
    regimes = sorted(df["regime"].unique())
    for reg in regimes:
        sub = df[df["regime"] == reg]
        n_all = len(sub)
        print(f"\n  [{reg.upper()}]  n={n_all}")
        print(f"  {'Hold':>5}  {'n':>5}  {'WR≥5%':>6}  {'WR≥0%':>6}  {'Avg%':>6}  {'Med%':>6}  {'Sharpe':>7}  {'SL%':>5}")
        print("  " + "-" * 52)
        for h in HOLD_PERIODS:
            s = _stats(sub[f"ret_t{h}"].dropna())
            print(f"  T+{h:<3}  {s['n']:>5}  {s['wr5']:>6.1f}  {s['wr0']:>6.1f}  "
                  f"{s['avg']:>+6.2f}  {s['med']:>+6.2f}  {s['sharpe']:>7.2f}  {s['sl_pct']:>5.1f}")

    # ── 11. By sector (T+20 vs T+25) ─────────────────────────────────────────
    print(f"\n{'─'*W}")
    print("  THEO SECTOR: T+20 vs T+25 (avg return %)")
    print(f"{'─'*W}")
    print(f"  {'Sector':<28}  {'n':>5}  {'T+20 avg':>9}  {'T+25 avg':>9}  {'Delta':>7}  {'Winner':>7}")
    print("  " + "-" * 68)
    sectors = df.groupby("sector")["ret_t20"].count().sort_values(ascending=False)
    for sec in sectors.index:
        sub = df[df["sector"] == sec]
        s20 = sub["ret_t20"].mean()
        s25 = sub["ret_t25"].mean() if "ret_t25" in sub else np.nan
        delta = s25 - s20 if not np.isnan(s25) else np.nan
        winner = "T+25 ↑" if delta > 0.3 else ("T+20 ↑" if delta < -0.3 else "≈ tie")
        print(f"  {sec:<28}  {len(sub):>5}  {s20:>+9.2f}  "
              f"{s25:>+9.2f}  {delta:>+7.2f}  {winner:>7}")

    # ── 12. Year-by-year breakdown ────────────────────────────────────────────
    print(f"\n{'─'*W}")
    print("  THEO NĂM")
    print(f"{'─'*W}")
    df["year"] = df["date"].dt.year
    for yr in sorted(df["year"].unique()):
        sub = df[df["year"] == yr]
        print(f"\n  [{yr}]  n={len(sub)}")
        print(f"  {'Hold':>5}  {'WR≥5%':>6}  {'WR≥0%':>6}  {'Avg%':>6}  {'Sharpe':>7}")
        print("  " + "-" * 38)
        for h in HOLD_PERIODS:
            s = _stats(sub[f"ret_t{h}"].dropna())
            if s["n"] == 0:
                continue
            print(f"  T+{h:<3}  {s['wr5']:>6.1f}  {s['wr0']:>6.1f}  "
                  f"{s['avg']:>+6.2f}  {s['sharpe']:>7.2f}")

    # ── 13. Return distribution analysis ──────────────────────────────────────
    print(f"\n{'─'*W}")
    print("  PHÂN PHỐI LỢI NHUẬN (percentiles)")
    print(f"{'─'*W}")
    pcts_labels = [5, 10, 25, 50, 75, 90, 95]
    hdr_pct = f"  {'Hold':>5}  " + "  ".join(f"p{p:>2}" for p in pcts_labels)
    print(hdr_pct)
    print("  " + "-" * (len(hdr_pct)))
    for h in HOLD_PERIODS:
        col   = f"ret_t{h}"
        pvals = df[col].dropna().quantile([p / 100 for p in pcts_labels])
        row_s = "  ".join(f"{v:>+5.1f}" for v in pvals)
        print(f"  T+{h:<3}  {row_s}")

    # ── 14. Excel ─────────────────────────────────────────────────────────────
    print(f"\nWriting Excel → {OUTPUT_XLS.name} …")

    # Build per-signal sheet
    display = df.copy()
    display["date"] = display["date"].dt.strftime("%d/%m/%Y")
    for h in HOLD_PERIODS:
        col = f"ret_t{h}"
        display[f"T+{h} %"] = display[col]
    display["Best Hold"] = display[[f"ret_t{h}" for h in HOLD_PERIODS]].idxmax(axis=1).str.replace("ret_t", "T+")
    display = display.drop(columns=[f"ret_t{h}" for h in HOLD_PERIODS])

    # Summary table
    sum_rows = []
    for h in HOLD_PERIODS:
        s = global_stats[h]
        sum_rows.append({
            "Hold Period": f"T+{h}",
            "n": s["n"],
            "WR ≥+5%": f"{s['wr5']:.1f}%",
            "WR ≥ 0%": f"{s['wr0']:.1f}%",
            "Avg Return": s["avg"],
            "Median Return": s["med"],
            "Sharpe": s["sharpe"],
            "P&L (M VND)": s["pnl_m"],
            "SL Rate (≤-10%)": f"{s['sl_pct']:.1f}%",
        })
    sum_df = pd.DataFrame(sum_rows)

    # Regime breakdown table
    reg_rows = []
    for reg in regimes:
        sub = df[df["regime"] == reg]
        for h in HOLD_PERIODS:
            s = _stats(sub[f"ret_t{h}"].dropna())
            reg_rows.append({
                "Regime": reg, "Hold": f"T+{h}",
                "n": s["n"], "WR≥5%": s["wr5"], "WR≥0%": s["wr0"],
                "Avg": s["avg"], "Sharpe": s["sharpe"], "SL%": s["sl_pct"],
            })
    reg_df = pd.DataFrame(reg_rows)

    # Sector breakdown
    sec_rows = []
    for sec in sectors.index:
        sub = df[df["sector"] == sec]
        row = {"Sector": sec, "n": len(sub)}
        for h in HOLD_PERIODS:
            s = _stats(sub[f"ret_t{h}"].dropna())
            row[f"T+{h} WR5%"] = s["wr5"]
            row[f"T+{h} Avg%"] = s["avg"]
            row[f"T+{h} Sharpe"] = s["sharpe"]
        row["Best Hold"] = max(HOLD_PERIODS,
            key=lambda h: _stats(sub[f"ret_t{h}"].dropna())["avg"])
        sec_rows.append(row)
    sec_df = pd.DataFrame(sec_rows)

    with pd.ExcelWriter(OUTPUT_XLS, engine="openpyxl") as writer:
        sum_df.to_excel(writer, sheet_name="📊 So sánh Hold Period", index=False)
        reg_df.to_excel(writer, sheet_name="🎯 Theo Regime",   index=False)
        sec_df.to_excel(writer, sheet_name="🏭 Theo Sector",   index=False)
        display.to_excel(writer, sheet_name="📋 Chi tiết Signal", index=False)

        # Formatting
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        HDR_FILL = PatternFill("solid", fgColor="1F4E79")
        BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
        F_WHITE  = Font(color="FFFFFF", bold=True)
        F_BOLD   = Font(bold=True)

        for ws in writer.sheets.values():
            for col in ws.columns:
                ml = max((len(str(c.value or "")) for c in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 30)
            for cell in ws[1]:
                if cell.value:
                    cell.fill = HDR_FILL
                    cell.font = F_WHITE
                    cell.alignment = Alignment(horizontal="center")

        # Highlight best hold period in summary sheet
        ws_sum = writer.sheets["📊 So sánh Hold Period"]
        sharpe_col = list(sum_df.columns).index("Sharpe") + 1
        avg_col    = list(sum_df.columns).index("Avg Return") + 1
        best_sharpe_row = sum_df["Sharpe"].idxmax() + 2
        best_avg_row    = sum_df["Avg Return"].idxmax() + 2
        for col_i in range(1, len(sum_df.columns) + 1):
            ws_sum.cell(best_sharpe_row, col_i).fill = BEST_FILL
            ws_sum.cell(best_sharpe_row, col_i).font = F_BOLD

    import shutil
    dest = Path.home() / "Desktop" / "compare_hold_period.xlsx"
    shutil.copy(OUTPUT_XLS, dest)
    print(f"\n✅ Done!  →  {dest}")


if __name__ == "__main__":
    run()
